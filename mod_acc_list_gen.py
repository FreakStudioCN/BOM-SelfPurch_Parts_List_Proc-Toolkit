# Python env   : Python 3.8+（需支持 pathlib、f-string 及 pandas/openpyxl 最新 API）
# -*- coding: utf-8 -*-
# @Time    : 2025/12/13 下午6:30
# @Author  : 李清水
# @File    : mod_acc_list_gen.py
# @Description : 遍历当前目录下符合规则的模块/扩展板文件夹，自动生成标准化配件清单Excel文件
#                核心功能：筛选目标文件夹→创建带预设列的Excel→设置公式（序号自增、总价计算）→优化格式（边框/居中/表头加粗）

import os
import re
import copy
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def gen_module_accessory_lists():
    root_dir = Path(os.getcwd())
    # -------------------------- 核心修改：筛选正则同时匹配“模块”或“扩展板” --------------------------
    # 匹配含“模块”或“扩展板”+ 以“-V版本号”结尾的文件夹
    module_folder_pattern = re.compile(r'.+(模块|扩展板)-V\d+\.\d+(\.\d+)?$')
    list_columns = ["No.", "Quantity", "Manufacturer Part", "Price", "Value", "淘宝链接", "下单配置", "最小起订量"]

    for folder in root_dir.iterdir():
        # 同时筛选“模块”或“扩展板”的文件夹，排除.idea
        if folder.is_dir() and module_folder_pattern.match(folder.name) and ".idea" not in folder.name:
            # 文件名格式：ModAcc_文件夹完整名字.xlsx
            list_filename = f"ModAcc_{folder.name}.xlsx"
            list_filepath = folder / list_filename

            # 初始化Excel
            df = pd.DataFrame(columns=list_columns)
            with pd.ExcelWriter(list_filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='配件清单', index=False)

            # 优化格式
            wb = load_workbook(list_filepath)
            ws = wb['配件清单']
            ws['A2'] = 1
            ws['A3'] = '=A2+1'
            ws['E2'] = '=B2*D2'
            ws['E3'] = '=B3*D3'

            # 边框+居中+表头加粗
            for row in range(1, 4):
                for col in range(1, len(list_columns)+1):
                    ws.cell(row=row, column=col).border = thin_border
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    if row == 1:
                        original_font = ws.cell(row=row, column=col).font
                        new_font = copy.copy(original_font)
                        new_font.bold = True
                        ws.cell(row=row, column=col).font = new_font

            # 列宽
            column_widths = [6, 10, 25, 10, 10, 30, 20, 12]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64+i)].width = width

            wb.save(list_filepath)
            print(f"✅ 生成：{list_filepath}")

    print("\n🎉 所有模块/扩展板的配件清单生成完成！")

if __name__ == "__main__":
    gen_module_accessory_lists()