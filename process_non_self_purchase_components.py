# Python env   : Python 3.8+（需支持 pathlib、f-string 及 pandas/openpyxl 最新 API）
# -*- coding: utf-8 -*-
# @Time    : 2025/12/13 下午6:25
# @Author  : 李清水
# @File    : process_non_self_purchase_components.py
# @Description : 处理BOM文件，提取非自采器件数据，实现分类（常规/特殊）、去重，最终生成格式化Excel汇总表
#                核心逻辑：查找BOM文件→清理无效数据→判定非自采器件→元器件分类→去重→生成样式化Excel

import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

def find_bom_files(root_dir):
    """
    查找同级文件夹下子文件夹中所有BOM/B0M开头的Excel文件
    :param root_dir: 根目录（当前脚本所在目录）
    :return: BOM文件路径列表
    """
    bom_files = []
    # 遍历根目录下的所有子文件夹
    for dirpath, dirnames, filenames in os.walk(root_dir):
        # 排除根目录本身，只处理子文件夹
        if dirpath == root_dir:
            continue
        for filename in filenames:
            # 匹配BOM/B0M开头的Excel文件（.xlsx/.xls）
            if re.match(r'^[Bb][0Oo][Mm]', filename, re.IGNORECASE) and filename.endswith(('.xlsx', '.xls')):
                bom_files.append(os.path.join(dirpath, filename))
    return bom_files


def clean_invalid_content(val):
    """
    清理无效内容（NaN、空白字符串、全角空格、制表符、换行符）
    :param val: 待清理的值
    :return: 清理后的值，无效则返回None
    """
    if pd.isna(val):
        return None
    # 转换为字符串并清理
    str_val = str(val).strip()
    # 移除全角空格、制表符、换行符
    str_val = str_val.replace('\u3000', '').replace('\t', '').replace('\n', '').replace('\r', '')
    return str_val if str_val else None


def match_column_name(df_columns, target_names):
    """
    匹配列名（支持变体）
    :param df_columns: DataFrame列名列表
    :param target_names: 目标列名列表（变体）
    :return: 匹配到的列名，无则返回None
    """
    df_cols_lower = [col.lower() for col in df_columns]
    for target in target_names:
        target_lower = target.lower()
        for idx, col_lower in enumerate(df_cols_lower):
            if target_lower in col_lower:
                return df_columns[idx]
    return None


def judge_non_self_purchase(df):
    """
    判定非自采器件（淘宝链接、下单配置、最小起订量均无有效内容）
    :param df: 原始BOM DataFrame
    :return: 非自采器件DataFrame
    """
    # 列名变体映射
    column_mapping = {
        'taobao_link': ['淘宝链接', '淘宝网址', 'taobao', 'taobao url'],
        'order_config': ['下单配置', '规格', '配置', 'spec', 'specification'],
        'min_order': ['最小起订量', '订购量', '最小订购量', 'moq', 'min order']
    }

    # 匹配列名
    taobao_col = match_column_name(df.columns, column_mapping['taobao_link'])
    order_col = match_column_name(df.columns, column_mapping['order_config'])
    min_order_col = match_column_name(df.columns, column_mapping['min_order'])

    # 复制DataFrame用于处理
    df_copy = df.copy()

    # 清理关键列内容
    if taobao_col:
        df_copy['clean_taobao'] = df_copy[taobao_col].apply(clean_invalid_content)
    else:
        df_copy['clean_taobao'] = None  # 未找到列视为无内容

    if order_col:
        df_copy['clean_order'] = df_copy[order_col].apply(clean_invalid_content)
    else:
        df_copy['clean_order'] = None

    if min_order_col:
        df_copy['clean_min_order'] = df_copy[min_order_col].apply(clean_invalid_content)
    else:
        df_copy['clean_min_order'] = None

    # 判定非自采器件（三列均无有效内容）
    non_self_purchase = df_copy[
        (df_copy['clean_taobao'].isna()) &
        (df_copy['clean_order'].isna()) &
        (df_copy['clean_min_order'].isna())
        ].copy()

    # 添加模块名称（从文件名提取）
    if 'file_path' in df_copy.columns:
        non_self_purchase['模块名称'] = non_self_purchase['file_path'].apply(
            lambda x: os.path.basename(x).replace('.xlsx', '').replace('.xls', '')
        )

    return non_self_purchase


def classify_component(designator):
    """
    按位号前缀分类元器件（大分类）
    :param designator: 位号（Designator）
    :return: 元器件大分类
    """
    if pd.isna(designator):
        return '未知（无位号）'

    designator_str = str(designator).strip().upper()

    # 大分类规则
    if designator_str.startswith('R'):
        return '电阻'
    elif designator_str.startswith('C'):
        return '电容'
    elif designator_str.startswith('LED'):
        return '二极管'
    elif designator_str.startswith('Q'):
        return '晶体管'
    elif designator_str.startswith('U'):
        return '集成电路'
    elif designator_str.startswith('SW'):
        return '开关'
    elif any(prefix in designator_str for prefix in ['J', 'CN', 'USB']):
        return '连接器'
    else:
        return '其他'


def process_bom_file(file_path):
    """
    处理单个BOM文件
    :param file_path: BOM文件路径
    :return: 处理后的非自采器件DataFrame
    """
    try:
        # 读取Excel文件（支持多sheet，取第一个sheet）
        df = pd.read_excel(file_path, sheet_name=0)
        df['file_path'] = file_path  # 记录文件路径

        # 1. 判定非自采器件
        non_self_df = judge_non_self_purchase(df)
        if non_self_df.empty:
            return None

        # 2. 匹配核心字段（处理列名变体）
        field_mapping = {
            'designator': ['designator', '位号', '元件位号'],
            'supplier_part': ['supplier part', '立创编号', '供应商编号', 'supplier p/n'],
            'manufacturer_part': ['manufacturer part', '器件型号', '型号', 'mfr p/n', 'manufacturer p/n'],
            'manufacturer': ['manufacturer', '制造商', '品牌']
        }

        # 匹配各字段列名
        designator_col = match_column_name(df.columns, field_mapping['designator']) or 'Designator'
        supplier_part_col = match_column_name(df.columns, field_mapping['supplier_part']) or 'Supplier Part'
        manufacturer_part_col = match_column_name(df.columns, field_mapping['manufacturer_part']) or 'Manufacturer Part'
        manufacturer_col = match_column_name(df.columns, field_mapping['manufacturer']) or 'Manufacturer'

        # 确保必要列存在（不存在则设为NaN）
        for col in [designator_col, supplier_part_col, manufacturer_part_col, manufacturer_col]:
            if col not in non_self_df.columns:
                non_self_df[col] = np.nan

        # 3. 分类元器件
        non_self_df['元器件类型'] = non_self_df[designator_col].apply(classify_component)

        # 4. 提取核心字段
        result_df = non_self_df[[
            '模块名称',
            designator_col,
            supplier_part_col,
            '元器件类型',
            manufacturer_part_col,
            manufacturer_col
        ]].copy()

        # 重命名为标准字段名
        result_df.columns = [
            '模块名称',
            'Designator',
            'Supplier Part',
            '元器件类型',
            'Manufacturer Part',
            'Manufacturer'
        ]

        # 清理字段内容
        for col in result_df.columns:
            result_df[col] = result_df[col].apply(clean_invalid_content)

        return result_df

    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {str(e)}")
        return None


def deduplicate_components(all_components):
    """
    按Supplier Part去重（保留首次出现记录）
    :param all_components: 所有非自采器件DataFrame
    :return: 去重后的DataFrame
    """
    if all_components.empty:
        return all_components

    # 按Supplier Part去重，保留首次出现
    deduplicated = all_components.drop_duplicates(subset=['Supplier Part'], keep='first').copy()
    return deduplicated


def split_regular_special(components_df):
    """
    区分常规器件和特殊器件
    特殊器件：缺失Designator、缺失Supplier Part、分类为"其他"
    :param components_df: 去重后的元器件DataFrame
    :return: 常规器件DataFrame、特殊器件DataFrame
    """
    # 判定特殊器件条件
    condition1 = components_df['Designator'].isna()  # 缺失位号
    condition2 = components_df['Supplier Part'].isna()  # 缺失Supplier Part
    condition3 = components_df['元器件类型'] == '其他'  # 分类为其他

    special_df = components_df[condition1 | condition2 | condition3].copy()
    regular_df = components_df[~(condition1 | condition2 | condition3)].copy()

    # 常规器件：保留核心3列（元器件类型、元器件名字、元器件编号）
    regular_final = regular_df[[
        '元器件类型',
        'Manufacturer Part',  # 元器件名字
        'Supplier Part'  # 元器件编号
    ]].copy()
    regular_final.columns = ['元器件类型', '元器件名字', '元器件编号']

    # 特殊器件：保留模块名称+核心3列
    special_final = special_df[[
        '模块名称',
        '元器件类型',
        'Manufacturer Part',
        'Supplier Part'
    ]].copy()
    special_final.columns = ['模块名称', '元器件类型', '元器件名字', '元器件编号']

    # 填充空值为"无"
    regular_final = regular_final.fillna('无')
    special_final = special_final.fillna('无')

    return regular_final, special_final


def format_excel_file(regular_df, special_df, output_path):
    """
    生成格式化Excel文件
    :param regular_df: 常规器件DataFrame
    :param special_df: 特殊器件DataFrame
    :param output_path: 输出文件路径
    """
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "非自采器件分类汇总"

    # 定义样式
    # 奇偶行填充色（浅灰#F5F5F5，白色#FFFFFF）
    gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    # 表头字体
    header_font = Font(bold=True)

    # 1. 写入常规器件
    current_row = 1
    # 常规器件表头
    regular_headers = ['元器件类型', '元器件名字', '元器件编号']
    for col, header in enumerate(regular_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font

    current_row += 1

    # 写入常规器件数据
    for idx, (_, row) in enumerate(regular_df.iterrows(), 1):
        for col, header in enumerate(regular_headers, 1):
            ws.cell(row=current_row, column=col, value=row[header])
        # 设置行颜色
        fill = gray_fill if idx % 2 == 0 else white_fill
        for col in range(1, len(regular_headers) + 1):
            ws.cell(row=current_row, column=col).fill = fill
        current_row += 1

    # 2. 写入特殊器件（标题+空行分隔）
    if not special_df.empty:
        # 空行
        current_row += 1
        # 特殊器件标题
        special_title = "特殊非自采器件（含模块信息）"
        ws.cell(row=current_row, column=1, value=special_title).font = header_font
        current_row += 1

        # 特殊器件表头
        special_headers = ['模块名称', '元器件类型', '元器件名字', '元器件编号']
        for col, header in enumerate(special_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font

        current_row += 1

        # 写入特殊器件数据
        for idx, (_, row) in enumerate(special_df.iterrows(), 1):
            for col, header in enumerate(special_headers, 1):
                ws.cell(row=current_row, column=col, value=row[header])
            # 设置行颜色
            fill = gray_fill if idx % 2 == 0 else white_fill
            for col in range(1, len(special_headers) + 1):
                ws.cell(row=current_row, column=col).fill = fill
            current_row += 1

    # 3. 自适应列宽
    def auto_adjust_column_width(ws, headers):
        for col, header in enumerate(headers, 1):
            # 计算列内容最大长度
            max_length = len(str(header))
            for row in range(2, ws.max_row + 1):
                cell_value = str(ws.cell(row=row, column=col).value)
                # 中文按2字符，英文/数字按1字符计算
                char_count = sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in cell_value)
                if char_count > max_length:
                    max_length = char_count
            # 预留10%边距
            adjusted_width = max_length * 1.1
            ws.column_dimensions[chr(64 + col)].width = adjusted_width

    # 调整常规器件列宽
    auto_adjust_column_width(ws, regular_headers)
    # 若有特殊器件，调整特殊器件列宽（取最大宽度）
    if not special_df.empty:
        auto_adjust_column_width(ws, special_headers)

    # 保存文件
    wb.save(output_path)
    print(f"Excel文件已生成: {output_path}")


def main():
    """
    主函数：处理所有BOM文件并生成汇总表
    """
    # 1. 配置参数
    root_dir = os.path.dirname(os.path.abspath(__file__))  # 当前脚本所在目录
    output_excel = os.path.join(root_dir, "非自采器件分类汇总表.xlsx")

    # 2. 查找所有BOM文件
    print("正在查找BOM文件...")
    bom_files = find_bom_files(root_dir)
    if not bom_files:
        print("未找到任何BOM/B0M开头的Excel文件")
        return
    print(f"找到 {len(bom_files)} 个BOM文件")

    # 3. 处理所有BOM文件
    all_components = []
    print("正在处理BOM文件...")
    for file in bom_files:
        print(f"处理: {os.path.basename(file)}")
        component_df = process_bom_file(file)
        if component_df is not None and not component_df.empty:
            all_components.append(component_df)

    if not all_components:
        print("未找到任何非自采器件")
        return

    # 合并所有数据
    combined_df = pd.concat(all_components, ignore_index=True)
    print(f"共找到 {len(combined_df)} 个非自采器件（去重前）")

    # 4. 去重
    deduplicated_df = deduplicate_components(combined_df)
    print(f"去重后剩余 {len(deduplicated_df)} 个非自采器件")

    # 5. 区分常规和特殊器件
    regular_df, special_df = split_regular_special(deduplicated_df)
    print(f"常规非自采器件: {len(regular_df)} 个")
    print(f"特殊非自采器件: {len(special_df)} 个")

    # 6. 生成格式化Excel
    format_excel_file(regular_df, special_df, output_excel)

    print("处理完成！")


if __name__ == "__main__":
    main()