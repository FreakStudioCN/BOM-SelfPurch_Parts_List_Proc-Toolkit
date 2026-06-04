# Python env   : Python 3.8+（需支持 pathlib、f-string 及 pandas/openpyxl 最新 API）
# -*- coding: utf-8 -*-        
# @Time    : 2025/12/13 下午6:21   
# @Author  : 李清水            
# @File    : modacc_accessory_processor.py       
# @Description :  处理 ModAcc 系列配件清单 Excel 文件，自动筛选自采配件数据，按模块汇总并计算总金额，
#                 生成 2 个格式化 Excel 表（模块汇总表 + 去重类型表），支持同一模块行颜色统一、不同模块颜色交替，同时输出数据统计信息

import os
import re
import copy
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Side, Border

def extract_and_format_accessory():
    """
    完整功能：处理ModAcc系列配件清单文件
    1. 匹配ModAcc_模块名-V版本号.xlsx 或 ModAcc_模块名_v版本号.xlsx格式文件（兼容大小写）
    2. 提取指定核心列数据，筛选有效配件
    3. 按模块汇总，同一模块行颜色统一、不同模块颜色交替
    4. 生成2个格式化Excel表+统计信息
    """
    # 1. 基础配置（100%匹配用户指定核心列+颜色规则）
    root_dir = Path(os.getcwd())
    # 关键修改：正则表达式支持 -V/-v 或 _V/_v 两种版本号前缀，兼容大小写
    # 匹配规则：ModAcc_xxx-V1.0.xlsx、ModAcc_xxx_v1.2.0.xlsx、modacc_xxx_V2.5.xlsx 等
    accessory_file_pattern = re.compile(r'^ModAcc_.+([-_]v)\d+\.\d+(\.\d+)?\.xlsx$', re.IGNORECASE)

    # 核心列：用户指定的纯英文+中文列（缺少则跳过文件）
    core_columns = [
        'No.',  # 序号列
        'Quantity',  # 数量列（纯英文）
        'Manufacturer Part',  # 配件名称列（纯英文）
        'Price',  # 单价列（纯英文）
        'Value',  # 配件总价列（纯英文，无需额外计算）
        '淘宝链接',  # 自采标识列
        '下单配置',  # 自采标识列
        '最小起订量'  # 自采标识列
    ]
    # 去重依据列：按“配件名称+单价+淘宝链接”去重，避免重复类型
    unique_type_cols = ['Manufacturer Part', 'Price', '淘宝链接']
    # 有效数据筛选列：含任意一列非空即视为需处理的自采配件
    filter_columns = ['淘宝链接', '下单配置', '最小起订量']

    # 样式配置：同一模块同色，不同模块交替（浅蓝→浅绿→浅黄循环）
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    colors = {
        'module_table': [  # 模块汇总表颜色（3种交替）
            PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),  # 浅蓝
            PatternFill(start_color="F0F8E6", end_color="F0F8E6", fill_type="solid"),  # 浅绿
            PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")  # 浅黄
        ],
        'type_table': [  # 去重类型表颜色（奇偶行交替）
            PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),  # 浅灰（偶数行）
            PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # 白色（奇数行）
        ]
    }

    # 2. 遍历目录，提取有效配件数据
    all_accessory_data = []
    print("🔍 开始搜索当前目录及子目录下的ModAcc配件清单文件...")
    for dir_path, _, file_names in os.walk(root_dir):
        for file_name in file_names:
            if accessory_file_pattern.match(file_name):
                accessory_path = Path(dir_path) / file_name
                print(f"\n📄 找到目标文件：{accessory_path}")

                try:
                    # 读取Excel文件：按字符串格式读取，避免数值自动转换导致丢失
                    df = pd.read_excel(accessory_path, dtype=str, header=0)
                    df.columns = df.columns.str.strip()  # 去除列名前后空格（兼容文件格式差异）

                    # 校验核心列：缺少则提示并跳过该文件
                    missing_cols = [col for col in core_columns if col not in df.columns]
                    if missing_cols:
                        print(f"❌ 跳过{file_name}：缺少核心列 → {missing_cols}")
                        continue

                    # 筛选有效数据：保留“淘宝链接/下单配置/最小起订量”任意非空的行
                    df_filtered = df.copy()
                    valid_mask = pd.Series(False, index=df_filtered.index)
                    for col in filter_columns:
                        # 排除空值和纯空格的行
                        col_mask = df_filtered[col].notna() & (df_filtered[col].str.strip() != '')
                        valid_mask = valid_mask | col_mask
                    df_filtered = df_filtered[valid_mask]

                    # 无有效数据时提示
                    if df_filtered.empty:
                        print(f"ℹ️ {file_name}：无有效自采配件数据")
                        continue

                    # 处理数值列：转换为数值类型（空值填充为0，用于后续汇总计算）
                    df_calc = df_filtered[core_columns].copy()
                    numeric_cols = ['Quantity', 'Price', 'Value']  # 需转换的数值列
                    for col in numeric_cols:
                        df_calc[col] = pd.to_numeric(df_calc[col], errors='coerce').fillna(0)

                    # 关键修改：提取模块名称，兼容 -V/-v 和 _V/_v 两种分隔符
                    # 匹配规则：同时处理 ModAcc_xxx-V1.0.xlsx 和 ModAcc_xxx_v1.2.0.xlsx 格式
                    module_name = re.sub(
                        r'^ModAcc_(.+)[-_]v\d+\.\d+(\.\d+)?\.xlsx$',
                        r'\1',
                        file_name,
                        flags=re.IGNORECASE
                    )
                    # 新增“模块名称”列（作为首列，便于后续按模块汇总）
                    df_with_module = df_calc.copy()
                    df_with_module.insert(0, '模块名称', module_name)
                    all_accessory_data.append(df_with_module)

                    print(f"✅ 成功提取：{len(df_with_module)}条有效配件数据")

                except Exception as e:
                    # 捕获处理过程中的异常（如文件损坏、权限不足等）
                    print(f"❌ 处理{file_name}失败：{str(e)}")
                    continue

    # 无任何有效数据时，退出程序并提示
    if not all_accessory_data:
        print("\n⚠️ 未找到任何有效ModAcc配件清单数据，程序退出")
        return

    # 3. 数据汇总与计算：按模块统计总金额
    print("\n📊 开始汇总所有配件数据...")
    # 合并所有文件的有效数据，按“模块名称+序号”排序（保持数据逻辑连贯）
    df_total = pd.concat(all_accessory_data, ignore_index=True).sort_values(by=['模块名称', 'No.'])
    # 按模块分组，计算每个模块的“配件总金额”（基于已有Value列求和）
    module_total = df_total.groupby('模块名称')['Value'].sum().reset_index()
    module_total.rename(columns={'Value': '模块配件总金额'}, inplace=True)
    # 合并模块总金额到主数据：每个模块的所有行都显示该模块总金额
    df_total = pd.merge(df_total, module_total, on='模块名称', how='left')
    # 调整列顺序：模块名称 → 模块总金额 → 原始核心列（提升可读性）
    col_order = ['模块名称', '模块配件总金额'] + core_columns
    df_total = df_total[col_order]

    # 4. 生成文件1：按模块汇总的配件表（同一模块同色，不同模块交替）
    file1_name = "1_按模块汇总的配件表.xlsx"
    file1_path = root_dir / file1_name
    print(f"\n📝 正在生成文件1：{file1_path}")
    # 写入Excel（不包含索引）
    with pd.ExcelWriter(file1_path, engine='openpyxl') as writer:
        df_total.to_excel(writer, sheet_name='模块配件汇总', index=False)

    # 美化文件1：合并模块单元格+统一颜色+样式优化
    wb1 = load_workbook(file1_path)
    ws1 = wb1['模块配件汇总']
    max_row1, max_col1 = ws1.max_row, ws1.max_column

    # 步骤1：合并相同模块的单元格（模块名称列+模块总金额列），并记录模块行范围
    module_ranges = []  # 存储每个模块的行区间：[(起始行, 结束行), ...]
    if max_row1 > 1:
        current_module = ws1['A2'].value  # 从第2行（首行是表头）开始
        start_row = 2
        for row in range(3, max_row1 + 1):
            if ws1[f'A{row}'].value != current_module:
                # 合并当前模块的单元格
                ws1.merge_cells(f'A{start_row}:A{row - 1}')  # 模块名称列（A列）
                ws1.merge_cells(f'B{start_row}:B{row - 1}')  # 模块总金额列（B列）
                module_ranges.append((start_row, row - 1))  # 记录当前模块行范围
                # 更新当前模块和起始行
                current_module = ws1[f'A{row}'].value
                start_row = row
        # 处理最后一个模块
        ws1.merge_cells(f'A{start_row}:A{max_row1}')
        ws1.merge_cells(f'B{start_row}:B{max_row1}')
        module_ranges.append((start_row, max_row1))

    # 步骤2：按模块行范围统一颜色（同一模块同色，不同模块交替）
    print("🎨 正在优化文件1样式：同一模块统一颜色...")
    color_idx = 0  # 颜色索引（循环使用module_table的3种颜色）
    for (module_start, module_end) in module_ranges:
        current_color = colors['module_table'][color_idx % len(colors['module_table'])]
        # 给当前模块的所有行应用颜色+边框+居中
        for row in range(module_start, module_end + 1):
            for col in range(1, max_col1 + 1):
                cell = ws1.cell(row=row, column=col)
                cell.fill = current_color
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        color_idx += 1  # 下一个模块切换颜色

    # 步骤3：表头样式优化（加粗+边框+居中）
    for col in range(1, max_col1 + 1):
        header_cell = ws1.cell(row=1, column=col)
        new_font = copy.copy(header_cell.font)
        new_font.bold = True
        header_cell.font = new_font  # 表头加粗
        header_cell.border = thin_border
        header_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 步骤4：自适应列宽（适配中文和长文本，避免内容截断）
    for col in range(1, max_col1 + 1):
        max_width = 0
        for row in range(1, max_row1 + 1):
            cell_val = str(ws1.cell(row=row, column=col).value or "")
            # 中文占2个字符宽度，英文/数字占1个字符
            width = sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in cell_val)
            max_width = max(max_width, width)
        # 预留10%的宽度余量，避免拥挤
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = max_width * 0.95

    # 保存文件1
    wb1.save(file1_path)
    print(f"✅ 文件1生成完成：{file1_path}")

    # 5. 生成文件2：去重后的配件类型表（奇偶行交替颜色）
    file2_name = "2_去重后的配件类型表.xlsx"
    file2_path = root_dir / file2_name
    print(f"\n📝 正在生成文件2：{file2_path}")

    # 按指定列去重：保留第一条重复数据
    df_unique = df_total[core_columns].drop_duplicates(subset=unique_type_cols, keep='first').reset_index(drop=True)
    # 重新生成序号：去重后序号连续（避免原序号断裂）
    df_unique['No.'] = range(1, len(df_unique) + 1)
    # 写入Excel
    with pd.ExcelWriter(file2_path, engine='openpyxl') as writer:
        df_unique.to_excel(writer, sheet_name='去重配件类型', index=False)

    # 美化文件2：奇偶行交替颜色+样式优化
    wb2 = load_workbook(file2_path)
    ws2 = wb2['去重配件类型']
    max_row2, max_col2 = ws2.max_row, ws2.max_column

    # 步骤1：奇偶行交替颜色（偶数行浅灰，奇数行白色）
    print("🎨 正在优化文件2样式：奇偶行交替颜色...")
    for row in range(2, max_row2 + 1):  # 从第2行开始（第1行是表头）
        color_idx = 0 if row % 2 == 0 else 1  # 偶数行→浅灰，奇数行→白色
        current_color = colors['type_table'][color_idx]
        for col in range(1, max_col2 + 1):
            cell = ws2.cell(row=row, column=col)
            cell.fill = current_color
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 步骤2：表头样式优化（同文件1）
    for col in range(1, max_col2 + 1):
        header_cell = ws2.cell(row=1, column=col)
        new_font = copy.copy(header_cell.font)
        new_font.bold = True
        header_cell.font = new_font
        header_cell.border = thin_border
        header_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 步骤3：自适应列宽（同文件1）
    for col in range(1, max_col2 + 1):
        max_width = 0
        for row in range(1, max_row2 + 1):
            cell_val = str(ws2.cell(row=row, column=col).value or "")
            width = sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in cell_val)
            max_width = max(max_width, width)
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = max_width * 0.95

    # 保存文件2
    wb2.save(file2_path)
    print(f"✅ 文件2生成完成：{file2_path}")

    # 6. 输出统计信息：清晰展示处理结果
    print("\n" + "=" * 50)
    print("📋 ModAcc配件清单处理结果统计")
    print("=" * 50)
    total_modules = df_total['模块名称'].nunique()  # 去重后的模块数量
    total_accessories = len(df_total)  # 原始有效配件总条数
    total_unique_types = len(df_unique)  # 去重后配件类型数
    total_amount = df_total['模块配件总金额'].unique().sum()  # 总金额（避免重复计算）
    print(f"🔹 涉及模块数量：{total_modules} 个")
    print(f"🔹 有效配件总条数：{total_accessories} 条")
    print(f"🔹 去重后配件类型：{total_unique_types} 种")
    print(f"🔹 配件总金额：{total_amount:.2f} 元")
    print("=" * 50)


# 程序入口：直接运行时执行
if __name__ == "__main__":
    print("=" * 60)
    print("📦 ModAcc配件清单处理工具（完整最终版）")
    print("=" * 60)
    extract_and_format_accessory()
    print("\n🎉 所有处理完成！")