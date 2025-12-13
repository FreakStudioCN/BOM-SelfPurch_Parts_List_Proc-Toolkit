# Python env   : Python 3.8+ï¼ˆéœ€æ”¯æŒ pathlibã€f-string åŠ pandas/openpyxl æœ€æ–° APIï¼‰
# -*- coding: utf-8 -*-
# @Time    : 2025/12/13 ä¸‹åˆ6:35
# @Author  : ææ¸…æ°´
# @File    : extract_bom_components.py
# @Description : å¤„ç†BOMæ–‡ä»¶ï¼ˆå‘½åæ ¼å¼ï¼šBOM_æ¨¡å—å-vç‰ˆæœ¬å·.xlsx/xlsï¼‰ï¼Œæå–éœ€è‡ªè¡Œé‡‡è´­çš„å…ƒå™¨ä»¶æ•°æ®
#                æ ¸å¿ƒåŠŸèƒ½ï¼šç­›é€‰è‡ªé‡‡æ•°æ®â†’æŒ‰æ¨¡å—æ±‡æ€»å¹¶è®¡ç®—æ€»ä»·â†’ç”Ÿæˆå¸¦æ ·å¼çš„Excelè¡¨ï¼ˆæ¨¡å—æ±‡æ€»è¡¨+å»é‡ç±»å‹è¡¨ï¼‰â†’è¾“å‡ºç»Ÿè®¡ä¿¡æ¯

import os
import re
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Side, Border

def extract_and_format_bom():
    root_dir = Path(os.getcwd())
    bom_file_pattern = re.compile(r'^BOM_.+-v\d+\.\d+(\.\d+)?\.(xlsx|xls)$', re.IGNORECASE)

    # æ ¸å¿ƒåˆ—ï¼ˆåŒ¹é…ä½ çš„BOMï¼‰
    core_columns = [
        'Manufacturer Part', 'Quantity', 'Designator', 'Supplier Part',
        'LCSC Price', 'Value', 'æ·˜å®é“¾æ¥', 'ä¸‹å•é…ç½®', 'æœ€å°èµ·è®¢é‡'
    ]
    unique_type_cols = ['Manufacturer Part', 'Supplier Part', 'LCSC Price', 'æ·˜å®é“¾æ¥']
    filter_columns = ['æ·˜å®é“¾æ¥', 'ä¸‹å•é…ç½®', 'æœ€å°èµ·è®¢é‡']

    # æ ·å¼é…ç½®ï¼šæ–°å¢ç»†è¾¹æ¡†
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    colors = {
        'module_table': [
            PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),
            PatternFill(start_color="F0F8E6", end_color="F0F8E6", fill_type="solid"),
            PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
        ],
        'type_table': [
            PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
            PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ]
    }

    # æå–æ•°æ®
    all_self_purchase = []
    print("ğŸ” æœç´¢BOMæ–‡ä»¶...")
    for dir_path, _, file_names in os.walk(root_dir):
        for file_name in file_names:
            if bom_file_pattern.match(file_name):
                bom_path = Path(dir_path) / file_name
                print(f"ğŸ“„ æ‰¾åˆ°ï¼š{bom_path}")
                try:
                    df = pd.read_excel(bom_path, dtype=str, header=0)
                    df.columns = df.columns.str.strip()

                    missing_cols = [col for col in core_columns if col not in df.columns]
                    if missing_cols:
                        print(f"âŒ è·³è¿‡{file_name}ï¼šç¼ºå°‘åˆ—{missing_cols}\n")
                        continue

                    # ç­›é€‰è‡ªè¡Œé‡‡è´­æ•°æ®
                    df_filtered = df.copy()
                    mask = pd.Series(False, index=df_filtered.index)
                    for col in filter_columns:
                        col_mask = df_filtered[col].notna() & (df_filtered[col].str.strip() != '')
                        mask = mask | col_mask
                    df_filtered = df_filtered[mask]

                    if df_filtered.empty:
                        print(f"â„¹ï¸ {file_name}æ— è‡ªé‡‡æ•°æ®\n")
                        continue

                    # å¤„ç†æ•°å€¼åˆ—
                    df_calc = df_filtered[core_columns].copy()
                    for col in ['Quantity', 'LCSC Price', 'Value']:
                        df_calc[col] = pd.to_numeric(df_calc[col], errors='coerce').fillna(0)

                    # æå–æ¨¡å—å
                    module_name = re.sub(r'^BOM_(.+)-v\d+\.\d+(\.\d+)?\.(xlsx|xls)$', r'\1', file_name, re.IGNORECASE)
                    df_with_module = df_calc.copy()
                    df_with_module.insert(0, 'æ¨¡å—åç§°', module_name)
                    all_self_purchase.append(df_with_module)
                    print(f"âœ… æå–{file_name}ï¼š{len(df_with_module)}ä¸ªå™¨ä»¶\n")

                except Exception as e:
                    print(f"âŒ å¤„ç†{file_name}å¤±è´¥ï¼š{str(e)}\n")
                    continue

    if not all_self_purchase:
        print("âš ï¸ æ— è‡ªé‡‡æ•°æ®")
        return
    df_with_module_all = pd.concat(all_self_purchase, ignore_index=True).sort_values(by='æ¨¡å—åç§°')

    # è®¡ç®—â€œè‡ªé‡‡å…ƒå™¨ä»¶æ€»ä»·â€
    module_total = df_with_module_all.groupby('æ¨¡å—åç§°')['Value'].sum().reset_index()
    module_total.rename(columns={'Value': 'è‡ªé‡‡å…ƒå™¨ä»¶æ€»ä»·'}, inplace=True)  # åˆ—åä¿®æ”¹
    df_with_module_all = pd.merge(df_with_module_all, module_total, on='æ¨¡å—åç§°', how='left')
    cols = df_with_module_all.columns.tolist()
    cols.insert(1, cols.pop(cols.index('è‡ªé‡‡å…ƒå™¨ä»¶æ€»ä»·')))
    df_with_module_all = df_with_module_all[cols]

    # ç”Ÿæˆæ–‡ä»¶1ï¼šæŒ‰æ¨¡å—æ±‡æ€»ï¼ˆå¸¦è¾¹æ¡†+æ–°åˆ—åï¼‰
    file1_path = root_dir / "1_æŒ‰æ¨¡å—æ±‡æ€»_è‡ªé‡‡å…ƒå™¨ä»¶.xlsx"
    with pd.ExcelWriter(file1_path, engine='openpyxl') as writer:
        df_with_module_all.to_excel(writer, sheet_name='æŒ‰æ¨¡å—æ±‡æ€»', index=False)

    wb1 = load_workbook(file1_path)
    ws1 = wb1['æŒ‰æ¨¡å—æ±‡æ€»']
    max_row1, max_col1 = ws1.max_row, ws1.max_column

    # åˆå¹¶å•å…ƒæ ¼ï¼ˆæ¨¡å—å+è‡ªé‡‡å…ƒå™¨ä»¶æ€»ä»·ï¼‰
    print("ğŸ“Š åˆå¹¶å•å…ƒæ ¼...")
    module_ranges = []
    if max_row1 > 1:
        current_module = ws1['A2'].value
        start_row = 2
        for row in range(3, max_row1 + 1):
            if ws1[f'A{row}'].value != current_module:
                ws1.merge_cells(f'A{start_row}:A{row - 1}')
                ws1.merge_cells(f'B{start_row}:B{row - 1}')  # ç¬¬2åˆ—æ˜¯æ–°åˆ—å
                module_ranges.append((start_row, row - 1))
                current_module = ws1[f'A{row}'].value
                start_row = row
        ws1.merge_cells(f'A{start_row}:A{max_row1}')
        ws1.merge_cells(f'B{start_row}:B{max_row1}')
        module_ranges.append((start_row, max_row1))

    # è®¾ç½®èƒŒæ™¯è‰²+è¾¹æ¡†+å±…ä¸­
    print("ğŸ¨ è®¾ç½®æ ·å¼...")
    color_idx = 0
    for (start_row, end_row) in module_ranges:
        current_color = colors['module_table'][color_idx % len(colors['module_table'])]
        for row in range(start_row, end_row + 1):
            for col in range(1, max_col1 + 1):
                ws1.cell(row=row, column=col).fill = current_color
                ws1.cell(row=row, column=col).border = thin_border  # æ·»åŠ è¾¹æ¡†
                ws1.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        color_idx += 1

    # è¡¨å¤´æ ·å¼ï¼ˆè¡¥å…¨è¾¹æ¡†+å±…ä¸­ï¼‰
    for col in range(1, max_col1 + 1):
        ws1.cell(row=1, column=col).border = thin_border
        ws1.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')

    # è‡ªé€‚åº”åˆ—å®½
    for col in range(1, max_col1 + 1):
        max_width = 0
        for row in range(1, max_row1 + 1):
            cell_val = str(ws1.cell(row=row, column=col).value or "")
            max_width = max(max_width, sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in cell_val))
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = max_width * 0.9

    wb1.save(file1_path)
    print(f"âœ… æ–‡ä»¶1ç”Ÿæˆï¼š{file1_path}\n")

    # ç”Ÿæˆæ–‡ä»¶2ï¼šå»é‡ç±»å‹ï¼ˆå¸¦è¾¹æ¡†ï¼‰
    df_type_unique = df_with_module_all[core_columns].drop_duplicates(subset=unique_type_cols,
                                                                      keep='first').reset_index(drop=True)
    file2_path = root_dir / "2_å»é‡_è‡ªé‡‡å…ƒå™¨ä»¶ç±»å‹.xlsx"
    with pd.ExcelWriter(file2_path, engine='openpyxl') as writer:
        df_type_unique.to_excel(writer, sheet_name='ç±»å‹æ±‡æ€»', index=False)

    wb2 = load_workbook(file2_path)
    ws2 = wb2['ç±»å‹æ±‡æ€»']
    max_row2, max_col2 = ws2.max_row, ws2.max_column

    # å¥‡å¶è¡ŒèƒŒæ™¯è‰²+è¾¹æ¡†+å±…ä¸­
    for row in range(2, max_row2 + 1):
        color_idx = 0 if row % 2 == 0 else 1
        current_color = colors['type_table'][color_idx]
        for col in range(1, max_col2 + 1):
            ws2.cell(row=row, column=col).fill = current_color
            ws2.cell(row=row, column=col).border = thin_border
            ws2.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

    # è¡¨å¤´æ ·å¼
    for col in range(1, max_col2 + 1):
        ws2.cell(row=1, column=col).border = thin_border
        ws2.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')

    # è‡ªé€‚åº”åˆ—å®½
    for col in range(1, max_col2 + 1):
        max_width = 0
        for row in range(1, max_row2 + 1):
            cell_val = str(ws2.cell(row=row, column=col).value or "")
            max_width = max(max_width, sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in cell_val))
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = max_width * 0.9

    wb2.save(file2_path)
    print(f"âœ… æ–‡ä»¶2ç”Ÿæˆï¼š{file2_path}\n")

    # ç»Ÿè®¡
    total = df_with_module_all.drop_duplicates(subset=['æ¨¡å—åç§°'])['è‡ªé‡‡å…ƒå™¨ä»¶æ€»ä»·'].sum()
    print("ğŸ“‹ ç»Ÿè®¡ï¼š")
    print(f"   - è‡ªé‡‡å™¨ä»¶æ•°ï¼š{len(df_with_module_all)}ä¸ª")
    print(f"   - è‡ªé‡‡æ€»é‡‘é¢ï¼š{total:.4f}å…ƒ")
    print(f"   - å»é‡ç±»å‹æ•°ï¼š{len(df_type_unique)}ç§")


if __name__ == "__main__":
    extract_and_format_bom()